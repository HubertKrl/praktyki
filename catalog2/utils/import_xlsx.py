import io
from openpyxl import load_workbook
from catalog2.models import Glassone, Glasstwo, Glassthree, Szyba, Ramka


def import_xlsx(file_obj):
    """
    Importuje dane z pliku XLSX do bazy danych Django.
    Plik musi posiadać trzy arkusze:
      - "Glassone"   z nagłówkami: ["ID", "Szyba1", "Grubość szyby1", "Ufactor", "Gfactor", "rw"]
      - "Glasstwo"   z nagłówkami: ["ID", "Szyba1", "Ramka1", "Szyba2", "Grubość całkowita", "Ufactor", "Gfactor", "rw"]
      - "Glassthree" z nagłówkami: ["ID", "Szyba1", "Ramka1", "Szyba2", "Ramka2", "Szyba3", "Grubość całkowita", "Ufactor", "Gfactor", "rw"]
    Jeśli rekord już istnieje (sprawdzany według kluczowych pól), to zostanie pominięty.
    """
    wb = load_workbook(filename=file_obj, data_only=True)

    # ------------------------------
    # Import arkusza "Glassone"
    # ------------------------------
    ws1 = wb["Glassone"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        # Rozpakowujemy dane:
        # [ID, Szyba1, Grubość szyby1, Ufactor, Gfactor, rw]
        _, szyba1_nazwa, szyba1_grubosc, ufactor, gfactor, rw = row

        # Pobieramy lub tworzymy obiekt Szyba
        szyba_obj, _ = Szyba.objects.get_or_create(
            nazwa=szyba1_nazwa,
            defaults={'grubosc': szyba1_grubosc}
        )

        # Sprawdzamy, czy rekord z tymi parametrami już istnieje w Glassone
        if Glassone.objects.filter(
                szyba1=szyba_obj,
                ufactor=ufactor,
                gfactor=gfactor,
                rw=rw
        ).exists():
            # Rekord istnieje – pomijamy
            continue

        # Rekord nie istnieje – tworzymy go
        Glassone.objects.create(
            szyba1=szyba_obj,
            ufactor=ufactor,
            gfactor=gfactor,
            rw=rw
        )

    # ------------------------------
    # Import arkusza "Glasstwo"
    # ------------------------------
    ws2 = wb["Glasstwo"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        # Rozpakowujemy dane:
        # [ID, Szyba1, Ramka1, Szyba2, Grubość całkowita, Ufactor, Gfactor, rw]
        _, szyba1_nazwa, ramka1_nazwa, szyba2_nazwa, grubosc_calkowita, ufactor, gfactor, rw = row

        # Pobieramy lub tworzymy powiązane obiekty
        szyba1_obj, _ = Szyba.objects.get_or_create(nazwa=szyba1_nazwa)
        ramka1_obj, _ = Ramka.objects.get_or_create(nazwa=ramka1_nazwa)
        szyba2_obj, _ = Szyba.objects.get_or_create(nazwa=szyba2_nazwa)

        if Glasstwo.objects.filter(
                szyba1=szyba1_obj,
                ramka1=ramka1_obj,
                szyba2=szyba2_obj,
                ufactor=ufactor,
                gfactor=gfactor,
                rw=rw
        ).exists():
            continue

        Glasstwo.objects.create(
            szyba1=szyba1_obj,
            ramka1=ramka1_obj,
            szyba2=szyba2_obj,
            ufactor=ufactor,
            gfactor=gfactor,
            rw=rw
        )

    # ------------------------------
    # Import arkusza "Glassthree"
    # ------------------------------
    ws3 = wb["Glassthree"]
    for row in ws3.iter_rows(min_row=2, values_only=True):
        # Rozpakowujemy dane:
        # [ID, Szyba1, Ramka1, Szyba2, Ramka2, Szyba3, Grubość całkowita, Ufactor, Gfactor, rw]
        (_, szyba1_nazwa, ramka1_nazwa,
         szyba2_nazwa, ramka2_nazwa,
         szyba3_nazwa, grubosc_calkowita, ufactor, gfactor, rw) = row

        szyba1_obj, _ = Szyba.objects.get_or_create(nazwa=szyba1_nazwa)
        ramka1_obj, _ = Ramka.objects.get_or_create(nazwa=ramka1_nazwa)
        szyba2_obj, _ = Szyba.objects.get_or_create(nazwa=szyba2_nazwa)
        ramka2_obj, _ = Ramka.objects.get_or_create(nazwa=ramka2_nazwa)
        szyba3_obj, _ = Szyba.objects.get_or_create(nazwa=szyba3_nazwa)

        if Glassthree.objects.filter(
                szyba1=szyba1_obj,
                ramka1=ramka1_obj,
                szyba2=szyba2_obj,
                ramka2=ramka2_obj,
                szyba3=szyba3_obj,
                ufactor=ufactor,
                gfactor=gfactor,
                rw=rw
        ).exists():
            continue

        Glassthree.objects.create(
            szyba1=szyba1_obj,
            ramka1=ramka1_obj,
            szyba2=szyba2_obj,
            ramka2=ramka2_obj,
            szyba3=szyba3_obj,
            ufactor=ufactor,
            gfactor=gfactor,
            rw=rw
        )

    return True