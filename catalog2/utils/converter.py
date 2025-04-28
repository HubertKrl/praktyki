import io
from openpyxl import Workbook
from catalog2.models import Glassone, Glasstwo, Glassthree


def generate_xlsx():
    wb = Workbook()

    # Arkusz 1: Glassone (pakiety jednoszybowe)
    ws1 = wb.active
    ws1.title = "Glassone"
    ws1.append(["ID", "Szyba1", "Grubość szyby1", "Ufactor", "Gfactor", "rw"])
    for obj in Glassone.objects.all():
        # Zakładamy, że pole 'szyba1' ma właściwości 'nazwa' oraz 'grubosc'
        ws1.append([
            obj.id,
            obj.szyba1.nazwa,
            obj.szyba1.grubosc,
            obj.ufactor,
            obj.gfactor,
            obj.rw
        ])

    # Arkusz 2: Glasstwo (pakiety dwuszybowe)
    ws2 = wb.create_sheet(title="Glasstwo")
    ws2.append(["ID", "Szyba1", "Ramka1", "Szyba2", "Grubość całkowita", "Ufactor", "Gfactor", "rw"])
    for obj in Glasstwo.objects.all():
        # W modelu Glasstwo metoda grubosc_calkowita jest zwykłą metodą (więc ją wywołujemy)
        ws2.append([
            obj.id,
            obj.szyba1.nazwa,
            obj.ramka1.nazwa,
            obj.szyba2.nazwa,
            obj.grubosc_calkowita(),  # wywołanie metody
            obj.ufactor,
            obj.gfactor,
            obj.rw
        ])

    # Arkusz 3: Glassthree (pakiety trzyszybowe)
    ws3 = wb.create_sheet(title="Glassthree")
    ws3.append(
        ["ID", "Szyba1", "Ramka1", "Szyba2", "Ramka2", "Szyba3", "Grubość całkowita", "Ufactor", "Gfactor", "rw"])
    for obj in Glassthree.objects.all():
        # W modelu Glassthree grubosc_calkowita jest dekorowana jako property, więc odwołujemy się bez nawiasów
        ws3.append([
            obj.id,
            obj.szyba1.nazwa,
            obj.ramka1.nazwa,
            obj.szyba2.nazwa,
            obj.ramka2.nazwa,
            obj.szyba3.nazwa,
            obj.grubosc_calkowita,  # property
            obj.ufactor,
            obj.gfactor,
            obj.rw
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output